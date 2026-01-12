import argparse, time, re, html, unicodedata, zipfile
from urllib.parse import quote_plus, urljoin
from pathlib import Path
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook, Workbook
import csv, io, subprocess, tempfile, requests
from PIL import Image
from urllib.parse import urlsplit, urlunsplit
import os
import hashlib

# =========================
# Config / constants
# =========================

ALLOW_GIF = False

PROFILE_DIR = ".mw_profile"
STATE_FILE  = "makerworld_state.json"

MAX_SHOPEE_IMG_BYTES = 2_000_000   # 2 MB
MAX_IMG_SIDE = 1200  

STEALTH_JS = r"""() => {
  Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
  const q=navigator.permissions&&navigator.permissions.query;
  if(q){navigator.permissions.query = p=>p.name==='notifications'
    ? Promise.resolve({state: Notification.permission}) : q(p);}
  Object.defineProperty(navigator,'languages',{get:()=>['en-US','en']});
  Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3,4]});
  Object.defineProperty(navigator,'hardwareConcurrency',{get:()=>8});
  const gp=WebGLRenderingContext.prototype.getParameter;
  WebGLRenderingContext.prototype.getParameter=function(p){
    if(p===37445) return "Intel Inc."; if(p===37446) return "Intel Iris OpenGL Engine";
    return gp.call(this,p);
  };
  window.chrome={runtime:{}};
}"""

# =========================
# Helpers: text & SEO
# =========================

def shopee_safe_image_url(u: str) -> str:
    if not u:
        return ""
    sp = urlsplit(u)
    q = sp.query or ""
    q = re.sub(r"(?i)/format,webp\b", "", q)
    q = re.sub(r"(?i)(^|[&?])format=webp\b", r"\1format=png", q)
    if not q or q == "x-oss-process=":
        return urlunsplit((sp.scheme, sp.netloc, sp.path, "", ""))
    return urlunsplit((sp.scheme, sp.netloc, sp.path, q, ""))

def _shopee_safe_jpg_bytes(img: Image.Image, max_bytes=MAX_SHOPEE_IMG_BYTES) -> bytes:
    img = img.convert("RGB")

    for q in [92, 90, 88, 86, 84, 82, 80, 78, 76, 74, 72, 70, 68, 66, 64]:
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=q, optimize=True, progressive=True)
        b = buf.getvalue()
        if len(b) <= max_bytes:
            return b

    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=60, optimize=True, progressive=True)
    return buf.getvalue()

def download_convert_shopee_image(url: str, save_dir: str, sku: str, idx: int) -> str | None:
    try:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        img_bytes = r.content
    except Exception as e:
        print("[img] download failed:", url, e)
        return None

    try:
        im = Image.open(io.BytesIO(img_bytes))

        w, h = im.size
        m = max(w, h)
        if m > MAX_IMG_SIDE:
            scale = MAX_IMG_SIDE / float(m)
            new_w = max(1, int(round(w * scale)))
            new_h = max(1, int(round(h * scale)))
            im = im.resize((new_w, new_h), Image.LANCZOS)

        jpg_bytes = _shopee_safe_jpg_bytes(im, MAX_SHOPEE_IMG_BYTES)

        Path(save_dir).mkdir(parents=True, exist_ok=True)
        filename = f"{sku}_{idx}.jpg"
        out_path = str(Path(save_dir) / filename)
        Path(out_path).write_bytes(jpg_bytes)

        return out_path
    except Exception as e:
        print("[img] convert failed:", url, e)
        return None

def download_and_convert_image(url: str, save_dir: str) -> str | None:
    try:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        img_bytes = r.content
    except Exception:
        return None

    try:
        im = Image.open(io.BytesIO(img_bytes))
        im = im.convert("RGB")

        Path(save_dir).mkdir(parents=True, exist_ok=True)
        base = re.sub(r'[^A-Za-z0-9_-]+', '_', Path(url).stem)[:60] or "img"
        out_path = f"{save_dir}/{base}.jpg"

        im.save(out_path, "JPEG", quality=90, subsampling=0)
        return out_path
    except Exception:
        return None


def _prefer_raw_gif(u: str) -> str:
    sp = urlsplit(u)
    if ".gif" in sp.path.lower():
        return urlunsplit((sp.scheme, sp.netloc, sp.path, "", ""))
    return u

def _norm_text(s: str) -> str:
    s = str(s or "")
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    s = re.sub(r"[^\w\s]", "", s)
    return s

def looks_like_cf(html_text:str)->bool:
    h=(html_text or "").lower()
    return any(k in h for k in ["just a moment","verifying you are human","cloudflare","cf-ray","cf-chl"])

def seo_title(t:str)->str:
    t=re.sub(r"\s+"," ", (t or "").strip())
    return (t + " – Cable Winder Organizer, Portable, 3D Print")[:255]

_REMOVE_LINE_PATTERNS = [
    re.compile(r"^\s*print\s*profile\s*:.*$", re.I),
    re.compile(r"^\s*sumber\s*desain\s*:.*$", re.I),
    re.compile(r"^\s*design\s*source\s*:.*$", re.I),
    re.compile(r"^\s*source\s*:.*$", re.I),
    re.compile(r"\bFAQ\b\s*$", re.I),
]

def _clean_desc_text(desc: str) -> str:
    if not desc:
        return ""
    text = html.unescape(desc).replace("\r\n", "\n").replace("\r", "\n").strip()
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    cleaned = []
    for ln in lines:
        if any(p.search(ln) for p in _REMOVE_LINE_PATTERNS):
            continue
        cleaned.append(ln)
    return strip_urls("\n".join(cleaned))

def load_text_file(path: str | None) -> str | None:
    if not path:
        return None
    p = Path(path)
    if not p.exists():
        print(f"[Template] desc template tidak ditemukan: {path}")
        return None
    return p.read_text(encoding="utf-8")

def format_desc_with_template(scraped_desc: str, template_text: str | None) -> str:
    scraped_clean = _clean_desc_text(scraped_desc)

    if not template_text:
        return seo_desc(scraped_desc)

    out = template_text.replace("{scraping_description_result}", scraped_clean or "-")

    out = out.replace("\r\n", "\n").replace("\r", "\n").strip()
    out = strip_urls(out)
    return out[:3000]

def seo_desc(desc:str)->str:
    parts=[]
    base = _clean_desc_text(desc)
    if base:
        parts.append(base)

    parts += [
        "Kegunaan: penggulung kabel USB/Type-C/Lightning.",
        "Material: PLA/ABS (sesuai profil cetak)."
    ]
    final = "\n\n".join([p for p in parts if p]).strip()
    return final[:3000]

def auto_scroll(page):
    page.evaluate("() => window.scrollTo(0,0)")
    for _ in range(14):
        page.evaluate("() => window.scrollBy(0,900)")
        time.sleep(0.12)

def get_gallery_urls(page)->list[str]:
    g=page.query_selector(".photo_show")
    urls=[]
    if g:
        for th in g.query_selector_all("img, .swiper-slide img, picture img"):
            try:
                th.scroll_into_view_if_needed()
                th.click(timeout=150)
                time.sleep(0.05)
            except:
                pass

        for im in g.query_selector_all("img, picture img"):
            try:
                u = im.get_attribute("src") or im.get_attribute("data-src") or im.get_attribute("srcset")
                if not u:
                    continue
                if "," in u:
                    u = u.split(",")[-1].strip().split(" ")[0]
                if u.startswith("//"):
                    u = "https:" + u
                if u.startswith("/"):
                    u = urljoin("https://makerworld.com", u)
                low=u.lower()
                if any(b in low for b in ["avatar","logo","icon","/comment","emote","placeholder",".svg",".ico"]):
                    continue
                exts = [".jpg",".jpeg",".png",".webp",".avif"]
                if ALLOW_GIF:
                    exts.append(".gif")
                path_low = urlsplit(u).path.lower()
                if not any(path_low.endswith(ext) for ext in exts):
                    continue
                u2 = shopee_safe_image_url(u)
                if u2 and u2 not in urls:
                    urls.append(u2)
                if len(urls)==8:
                    break
            except:
                pass

    if not urls:
        try:
            og = page.query_selector("meta[property='og:image']")
            if og:
                u = (og.get_attribute("content") or "").strip()
                if u:
                    urls = [shopee_safe_image_url(u)]
        except:
            pass

    return urls[:8]

# =========================
# XLSX sanitizer & header mapping
# =========================

SHEETVIEWS_RE = re.compile(rb"<sheetViews[\s\S]*?</sheetViews>")
IMG_LABEL_RE = re.compile(
    r"^(foto\s+(sampul|utama|produk\s+\d+)|cover\s*image|item\s*image\s*\d+)$",
    re.I
)
URL_RE = re.compile(
    r"(?i)\b(?:https?://|www\.)\S+|\b\S+\.(?:com|net|org|id|io|co|me|xyz|gg|ly|ai|app|site|store|shop|link|pdf|zip)\b\S*"
)

HEADER_ALIASES = {
    "Kategori": {"kategori", "category", "product category", "category id"},
    "Nama Produk": {"nama produk", "product name", "name", "nama"},
    "Deskripsi Produk": {"deskripsi produk", "product description", "description"},
    "Foto Produk": {"foto produk", "images", "image urls", "product images", "photo", "photos", "gambar", "url gambar"},
    "Harga": {"harga", "price"},
    "Stok": {"stok", "stock", "quantity"},
    "SKU Induk": {"sku induk", "sku", "parent sku", "model sku"},
    "Berat (gram)": {"berat gram", "berat", "berat produk", "weight", "weight g", "weight gram"},
    "Panjang Paket (cm)": {"panjang paket cm", "panjang cm", "panjang", "length", "length cm"},
    "Lebar Paket (cm)": {"lebar paket cm", "lebar cm", "lebar", "width", "width cm"},
    "Tinggi Paket (cm)": {"tinggi paket cm", "tinggi cm", "tinggi", "height", "height cm"},
    "Masa Garansi": {"masa garansi", "garansi", "warranty period", "warranty", "warranty duration"},
}

def strip_urls(text: str) -> str:
    if not text:
        return ""
    text = URL_RE.sub("", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

def sanitize_xlsx(src_path:str)->str:
    src = Path(src_path)
    if not src.exists():
        raise FileNotFoundError(f"Template tidak ditemukan: {src_path}")
    dst = src.with_name(src.stem + "_sanitized.xlsx")
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                data = SHEETVIEWS_RE.sub(b"", data)
            zout.writestr(item, data)
    return str(dst)

def create_minimal_template(path:str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Kategori","Nama Produk","Deskripsi Produk",
               "Harga","Stok","SKU Induk",
               "Berat (gram)","Panjang Paket (cm)","Lebar Paket (cm)","Tinggi Paket (cm)",
               "Foto Produk", "Masa Garansi"])
    wb.save(path)
    return path

def _find_header_positions(ws, search_rows=50, search_cols=160):
    for start_r in range(1, search_rows+1):
        for h_rows in (1, 2, 3):
            merged = []
            for c in range(1, search_cols+1):
                parts=[]
                for rr in range(start_r, min(start_r+h_rows, search_rows+1)):
                    v = ws.cell(rr, c).value
                    if v is not None and str(v).strip():
                        parts.append(str(v))
                merged.append(" ".join(parts) if parts else "")
            normalized = [_norm_text(v) for v in merged]
            col_to_text = {i+1: t for i,t in enumerate(normalized) if t}

            found={}
            for target, aliases in HEADER_ALIASES.items():
                alias_norm = {_norm_text(a) for a in aliases}
                for c, txt in col_to_text.items():
                    if txt in alias_norm:
                        found[target]=c
                        break

            image_cols=[]
            for c, txt in col_to_text.items():
                if IMG_LABEL_RE.match(txt):
                    image_cols.append(c)

            if ("Kategori" in found and "Nama Produk" in found and
                ("Deskripsi Produk" in found or image_cols or "Foto Produk" in found)):
                return (start_r + h_rows - 1, found, sorted(image_cols))
    return (None, {}, [])

def write_rows_to_shopee_template(template_path:str, out_path:str, rows:list[dict], sheet_name: str | None = None):
    try:
        patched = sanitize_xlsx(template_path)
    except Exception as e:
        print("[Template] gagal sanitize:", e)
        patched = create_minimal_template("_fallback_template.xlsx")

    try:
        wb = load_workbook(patched)
    except Exception as e:
        print("[openpyxl] gagal load:", e)
        patched = create_minimal_template("_fallback_template.xlsx")
        wb = load_workbook(patched)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb["Template"] if "Template" in wb.sheetnames else wb[wb.sheetnames[0]]

    hdr_row, colmap_found, image_cols = _find_header_positions(ws)
    if not hdr_row:
        print("[Header] tidak ketemu header valid. Buat sheet 'Template' minimal.")
        if "Template" in wb.sheetnames:
            del wb["Template"]
        ws = wb.create_sheet("Template")
        ws.append(["Kategori","Nama Produk","Deskripsi Produk",
                   "Harga","Stok","SKU Induk",
                   "Berat (gram)","Panjang Paket (cm)","Lebar Paket (cm)","Tinggi Paket (cm)",
                   "Foto Produk", "Masa Garansi"])
        hdr_row = 1
        colmap_found = {h: i+1 for i, h in enumerate(["Kategori","Nama Produk","Deskripsi Produk",
                                                      "Harga","Stok","SKU Induk",
                                                      "Berat (gram)","Panjang Paket (cm)","Lebar Paket (cm)","Tinggi Paket (cm)",
                                                      "Foto Produk", "Masa Garansi"])}
        image_cols = []

    def ensure_col(label):
        if label in colmap_found and colmap_found[label]:
            return colmap_found[label]
        new_idx = ws.max_column + 1
        ws.cell(hdr_row, new_idx).value = label
        colmap_found[label] = new_idx
        return new_idx

    for req in ["Kategori","Nama Produk","Deskripsi Produk"]:
        ensure_col(req)
    for maybe in ["Harga","Stok","SKU Induk","Berat (gram)","Panjang Paket (cm)","Lebar Paket (cm)","Tinggi Paket (cm)","Foto Produk", "Masa Garansi"]:
        ensure_col(maybe)

    img_cols=[]
    for c in range(1, ws.max_column+1):
        lab = _norm_text(ws.cell(hdr_row, c).value)
        if IMG_LABEL_RE.match(lab):
            img_cols.append(c)
    img_cols = sorted(img_cols)

    def _img_rank(c):
        t = _norm_text(ws.cell(hdr_row, c).value)
        if t in ("cover image", "foto sampul", "foto utama"):
            return (0, 0, c)
        m = re.search(r"(?:item image|foto produk)\s*(\d+)$", t)
        if m:
            return (1, int(m.group(1)), c)
        return (2, 999, c)

    img_cols = sorted(img_cols, key=_img_rank)

    def row_has_data(rr):
        for c in range(1, ws.max_column+1):
            if str(ws.cell(rr, c).value or "").strip():
                return True
        return False
    r = hdr_row + 1
    while row_has_data(r): r += 1

    def C(label): return colmap_found[label]

    for row in rows:
        ws.cell(r, C("Kategori")).value          = str(row["category_id"])
        ws.cell(r, C("Nama Produk")).value       = row["name"]
        ws.cell(r, C("Deskripsi Produk")).value  = row["description"]

        if "price" in row and "Harga" in colmap_found:
            ws.cell(r, C("Harga")).value = int(row["price"])
        if "stock" in row and "Stok" in colmap_found:
            ws.cell(r, C("Stok")).value  = int(row["stock"])
        if "sku" in row and "SKU Induk" in colmap_found:
            ws.cell(r, C("SKU Induk")).value = row["sku"]

        if "weight_kg" in row:
            if "Berat (gram)" in colmap_found:
                ws.cell(r, C("Berat (gram)")).value = int(round(row["weight_kg"]*1000))
        if "dims_cm" in row:
            L,W,H = row["dims_cm"]
            if "Panjang Paket (cm)" in colmap_found: ws.cell(r, C("Panjang Paket (cm)")).value = int(L)
            if "Lebar Paket (cm)"   in colmap_found: ws.cell(r, C("Lebar Paket (cm)")).value   = int(W)
            if "Tinggi Paket (cm)"  in colmap_found: ws.cell(r, C("Tinggi Paket (cm)")).value  = int(H)

        urls = row.get("image_urls", [])[:8]
        if img_cols:
            for idx, c in enumerate(img_cols):
                if idx < len(urls):
                    ws.cell(r, c).value = urls[idx]
        else:
            if "Foto Produk" in colmap_found:
                ws.cell(r, C("Foto Produk")).value = ",".join(urls)

        warranty_text = (row.get("warranty") or "No Warranty")
        if "Masa Garansi" in colmap_found:
            ws.cell(r, C("Masa Garansi")).value = warranty_text
        else:
            ensure_col("Masa Garansi")
            ws.cell(r, C("Masa Garansi")).value = warranty_text

        r += 1

    wb.save(out_path)
    print("[Shopee] file jadi →", out_path)

# =========================
# Scraping MakerWorld
# =========================

def scrape(keyword:str, max_results:int, headless:bool, proxy:str|None):
    results=[]
    with sync_playwright() as p:
        for engine in ("chromium","webkit","firefox"):
            ctx=None
            try:
                bt=getattr(p, engine)
                kwargs={"headless": headless, "args":["--no-sandbox","--disable-blink-features=AutomationControlled"]}
                if proxy: kwargs["proxy"]={"server":proxy}
                ctx=bt.launch_persistent_context(
                    user_data_dir=PROFILE_DIR,
                    **kwargs, locale="en-US", timezone_id="Asia/Jakarta",
                    viewport={"width":1400,"height":900}, device_scale_factor=2
                )
                ctx.add_init_script(STEALTH_JS)
                page=ctx.new_page(); page.set_default_navigation_timeout(120000)

                url=f"https://makerworld.com/en/search/models?keyword={quote_plus(keyword)}"
                print(f"[{engine}] open", url)
                page.goto(url, wait_until="domcontentloaded")
                try: page.wait_for_load_state("networkidle", timeout=15000)
                except: pass

                if looks_like_cf(page.content()):
                    print(f"[{engine}] Cloudflare detected")
                    if headless and not proxy:
                        try: ctx.storage_state(path=STATE_FILE)
                        except: pass
                        ctx.close()
                        return scrape(keyword, max_results, headless=False, proxy=proxy)
                    time.sleep(8)

                links=[]; seen=set()
                for a in page.query_selector_all("a[href*='/models/'], a[href*='/model/']"):
                    h=a.get_attribute("href")
                    if not h: continue
                    if h.startswith("/"): h=urljoin("https://makerworld.com", h)
                    if h not in seen: seen.add(h); links.append(h)
                    if len(links)>=max_results: break
                print(f"[{engine}] links:", len(links))

                for link in links[:max_results]:
                    page.goto(link, wait_until="domcontentloaded")
                    try: page.wait_for_load_state("networkidle", timeout=12000)
                    except: pass

                    title=""
                    try:
                        h1=page.query_selector("h1")
                        if h1: title=h1.inner_text().strip()
                    except: pass
                    if not title:
                        try:
                            og=page.query_selector("meta[property='og:title']")
                            if og: title=(og.get_attribute("content") or "").strip()
                        except: pass

                    desc=""
                    try:
                        md=page.query_selector("meta[name='description']")
                        if md: desc=(md.get_attribute("content") or "").strip()
                    except: pass

                    auto_scroll(page)
                    gal=get_gallery_urls(page)

                    results.append({
                        "title": title or link,
                        "description": desc,
                        "url": link,
                        "image_urls": gal
                    })
                try: ctx.storage_state(path=STATE_FILE)
                except: pass
                ctx.close()
                if results: break
            except Exception as e:
                print(f"[{engine}] error:", e)
                try:
                    if ctx: ctx.close()
                except: pass
    return results


# =========================
# GIF / Video helpers + Metadata writer
# =========================
def _ensure_dir(p: str):
    Path(p).parent.mkdir(parents=True, exist_ok=True)

def gif_to_mp4_bytes(gif_bytes: bytes):
    """Convert GIF -> MP4; fallback ambil first frame JPG.
       Memastikan dimensi genap agar kompatibel yuv420p/libx264."""
    import tempfile, subprocess, io
    from pathlib import Path
    from PIL import Image

    with tempfile.TemporaryDirectory() as td:
        gif_path = Path(td) / "in.gif"
        mp4_path = Path(td) / "out.mp4"
        gif_path.write_bytes(gif_bytes)

        mp4_bytes = None
        vf = "fps=24,scale=trunc(iw/2)*2:trunc(ih/2)*2"
        cmd = [
            "ffmpeg", "-y",
            "-i", str(gif_path),
            "-movflags", "faststart",
            "-pix_fmt", "yuv420p",
            "-c:v", "libx264",
            "-vf", vf,
            str(mp4_path),
        ]
        try:
            p = subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            if mp4_path.exists() and mp4_path.stat().st_size > 0:
                mp4_bytes = mp4_path.read_bytes()
        except Exception as e:
            try:
                print("[ffmpeg] convert failed:", e)
                p = subprocess.run(cmd, capture_output=True, text=True)
                print("[ffmpeg] stderr:", (p.stderr or "")[:1000])
            except Exception:
                pass

        jpg_bytes = None
        try:
            im = Image.open(io.BytesIO(gif_bytes))
            im.seek(0)
            rgb = im.convert("RGB")
            buf = io.BytesIO()
            rgb.save(buf, format="JPEG", quality=90)
            jpg_bytes = buf.getvalue()
        except Exception:
            pass

    return mp4_bytes, jpg_bytes

def handle_gif_url(url: str, save_dir="downloads"):
    """Download GIF & convert; return dict path hasil."""
    try:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        gif_bytes = r.content
    except Exception:
        return {"gif_path": None, "mp4_path": None, "jpg_path": None}

    _ensure_dir(save_dir + "/x")
    base = re.sub(r'[^A-Za-z0-9_-]+', '_', Path(url).stem)[:60] or "gif"
    gif_path = f"{save_dir}/{base}.gif"
    Path(gif_path).write_bytes(gif_bytes)

    mp4_bytes, jpg_bytes = gif_to_mp4_bytes(gif_bytes)
    mp4_path = None
    jpg_path = None
    if mp4_bytes:
        mp4_path = f"{save_dir}/{base}.mp4"
        Path(mp4_path).write_bytes(mp4_bytes)
    if jpg_bytes:
        jpg_path = f"{save_dir}/{base}.jpg"
        Path(jpg_path).write_bytes(jpg_bytes)

    return {"gif_path": gif_path, "mp4_path": mp4_path, "jpg_path": jpg_path}

def write_metadata_csv(path_csv: str, items: list[dict]):
    Path(path_csv).parent.mkdir(parents=True, exist_ok=True)
    with open(path_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["SKU","Nama Produk","MakerWorld Link","Image URLs","Video MP4 (local)","GIF First Frame (local)"])
        for it in items:
            w.writerow([
                it.get("sku",""),
                it.get("title",""),
                it.get("makerworld_url",""),
                " | ".join(it.get("image_urls", [])),
                it.get("video_path","") or "",
                it.get("gif_first_frame","") or "",
            ])
    print(f"✅ Metadata CSV tersimpan: {path_csv}")

# =========================
# Main
# =========================

def main():
    ap=argparse.ArgumentParser()

    global ALLOW_GIF
    ap.add_argument("-k","--keyword", required=True)
    ap.add_argument("-m","--max", type=int, default=1)
    ap.add_argument("--template", required=True)
    ap.add_argument("-o","--out", default="shopee_ready.xlsx")
    ap.add_argument("--category-id", required=True)
    ap.add_argument("--brand", default="No Brand")
    ap.add_argument("--price", type=float, default=45000)
    ap.add_argument("--stock", type=int, default=20)
    ap.add_argument("--weight-kg", type=float, default=0.15)
    ap.add_argument("--dims", default="10,10,3")  # L,W,H
    ap.add_argument("--sku-prefix", dest="sku_prefix", default="MW")
    ap.add_argument("--proxy", default=None)
    ap.add_argument("--headless", action="store_true")
    ap.add_argument("--sheet", default=None, help="Nama sheet bila template punya banyak tab")
    ap.add_argument("--warranty", default="Tidak bergaransi", help="Isi masa garansi jika tidak ada di template")

    ap.add_argument("--meta-out", default="metadata.csv", help="Output metadata CSV")
    ap.add_argument("--allow-gif", action="store_true", help="Aktifkan konversi GIF → MP4 (atau JPG frame pertama)")
    ap.add_argument("--download-dir", default="downloads", help="Folder simpan hasil unduhan")

    ap.add_argument("--desc-template", default=None, help="Path .txt template deskripsi, gunakan placeholder {scraping_description_result}")

    args=ap.parse_args()

    desc_template_text = load_text_file(args.desc_template)

    out_dir = os.path.dirname(args.out)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    global ALLOW_GIF
    ALLOW_GIF = bool(args.allow_gif)

    dims=tuple(int(x) for x in args.dims.split(",")[:3])

    print("[*] scraping …")
    recs=scrape(args.keyword, args.max, args.headless, args.proxy)
    if not recs:
        print("[!] Tidak ada hasil. Coba jalankan tanpa --headless atau pakai proxy residensial.")
        return

    rows=[]
    metas=[]
    
    for i, r in enumerate(recs, start=1):
        desc_final = format_desc_with_template(r["description"], desc_template_text)

        sku = f"{args.sku_prefix}-{i:04d}"

        urls_public = []
        for u in r.get("image_urls", [])[:8]:
            u2 = shopee_safe_image_url(u)
            if u2:
                urls_public.append(u2)
        urls_public = urls_public[:8]

        if not urls_public:
            print(f"[!] Gagal ambil gambar untuk SKU {sku}. Shopee butuh Cover image. Skip item ini.")
            continue

        rows.append({
            "category_id": args.category_id,
            "name": seo_title(r["title"]),
            "description": desc_final,
            "image_urls": urls_public,
            "price": args.price,
            "stock": args.stock,
            "weight_kg": args.weight_kg,
            "dims_cm": dims,
            "sku": sku,
        })

        meta = {
            "sku": sku,
            "title": seo_title(r["title"]),
            "makerworld_url": r.get("url", ""),
            "image_urls": r.get("image_urls", [])[:8],
            "video_path": "",
            "gif_first_frame": "",
        }

        if ALLOW_GIF:
            for u in meta["image_urls"]:
                u_try = _prefer_raw_gif(u)
                paths = handle_gif_url(u_try, save_dir=args.download_dir)
                if paths.get("mp4_path") or paths.get("jpg_path"):
                    meta["video_path"] = paths.get("mp4_path") or ""
                    meta["gif_first_frame"] = paths.get("jpg_path") or ""
                    break

        metas.append(meta)

    write_rows_to_shopee_template(args.template, args.out, rows, sheet_name=args.sheet)
    write_metadata_csv(args.meta_out, metas)
    print("✅ Selesai.")

if __name__ == "__main__":
    main()
